import datetime
from settings import month__, month, folder_path_machine1, folder_path_machine2
import calendar
import random
from openpyxl import load_workbook
import os
from pathlib import Path

month__.items()
# print(month__.items())
wb = load_workbook('import/ttnweight0.xlsx')
ws = wb.active


def examination_in_month_day(year, month_):  # проверка сколько дней в месяцу
    for key, value in month.items():
        if value == month_:
            return calendar.monthrange(int(year), int(key))[1]
    raise ValueError(f"Месяц {month_} не найде")


def date_number_document(colum):
    # Получаем данные из ячеек
    data_number = colum.cell(row=39, column=8).value
    data_number = int(str(data_number).strip('"'))  # Более безопасное преобразование
    data_month = colum.cell(row=39, column=13).value
    year = colum.cell(row=39, column=28).value
    year = str(year)[0:4]  # Преобразуем в строку и берем первые 4 символа

    # print(f"Начальные данные: число={data_number}, месяц={data_month}, год={year}")

    # count = 3
    current_day = data_number
    current_month = data_month
    current_year = int(year)
    # print(f"current_month: {type(current_month)} current_day: {type(current_day)}, current_year: {type(current_year)}")
    # for i in range(count):
    # Проверяем, не превышает ли текущий день количество дней в месяце
    days_in_month = examination_in_month_day(current_year, current_month)
    print(f"days in month: {type(days_in_month)}: {days_in_month}")
    if current_day > days_in_month:
        # Переходим к следующему месяцу
        current_day = 1
        #
        #     # Находим следующий месяц
        for key, value in month__.items():
            if value == current_month:
                next_key = int(key) % 12 + 1
                print(next_key)
                current_month = month[str(next_key)]
                print(current_month)

                # Если перешли на январь, увеличиваем год
                if next_key == 1:
                    current_year += 1
                break
    else:
        current_day += 1

    # Записываем данные в ячейки
    current_day = colum.cell(row=39, column=8, value=current_day).value
    current_month = colum.cell(row=39, column=13, value=current_month).value
    current_year = colum.cell(row=39, column=28, value=str(current_year)).value
    new_date = []

    for kay, value in month.items():
        if current_month == value:
            new_date.append(str(current_day))
            # print(new_data)
            new_date.append(kay)
            # print(new_data)
            new_date.append(str(current_year))
            # print(new_data)
            # print(new_data)
    new_date = ".".join(new_date)
    # new_date = colum.cell(row=5, column=103, value=new_date).value

    date = colum.cell(row=5, column=103).value
    # print(f"Дата добавлена {current_day}:{current_month}:{current_year}")
    return f"день: {current_day},  Месяц: {current_month}, год: {current_year} ----> Дата добавлена в ячейку r5c103 {new_date}, data= {type(date)}"


def examination_machine1(colum, row: int, column: int, basa):
    documen_namber = colum.cell(row=4, column=103).value
    documen_date = colum.cell(row=5, column=103).value
    document = f"ТТН {documen_namber} от {documen_date}"
    values = colum.cell(row=row, column=column).value
    if values == "Джанбулатов Руслан Джанболатович":
        ducument = colum.cell(row=62, column=31, value=document).value
        values = colum.cell(row=18, column=89).value
        values = colum.cell(row=26, column=105, value=values).value
        date_number_document(colum)

        return values, ducument, True
    else:
        values = colum.cell(row=18, column=89).value
        values = colum.cell(row=26, column=105, value=values).value
        machine1 = basa["machine1"]
        machine1_driver = machine1["driver"]
        machine1_driver = colum.cell(row=49, column=6, value=machine1_driver).value
        machine1_driver = colum.cell(row=33, column=94, value=machine1_driver).value
        machine1_driver = colum.cell(row=72, column=17, value=machine1_driver).value
        machine1_driver = colum.cell(row=69, column=73, value=machine1_driver).value
        machine1_certificate = machine1['certificate']
        machine1_certificate = colum.cell(row=49, column=51, value=machine1_certificate).value
        machine1_state_signs = machine1['state_signs']
        machine1_state_signs = colum.cell(row=45, column=63, value=machine1_state_signs).value
        machine1_trailer = machine1["trailer"]
        machine1_trailer = colum.cell(row=55, column=83, value=machine1_trailer).value
        ducument = colum.cell(row=62, column=31, value=document).value
        date_number_document(colum)

    return machine1_driver, machine1_certificate, machine1_state_signs, machine1_trailer, values, ducument


def examination_machine2(colum, row: int, column: int, base):
    document_number = colum.cell(row=4, column=103).value
    document_date = colum.cell(row=5, column=103).value
    document = f"ТТН {document_number} от {document_date}"
    print(document)
    values = colum.cell(row=row, column=column).value
    if values == "Гайирбеков Ибрагим Расулович":
        document = colum.cell(row=62, column=31, value=document).value
        values = colum.cell(row=18, column=89).value
        values = colum.cell(row=26, column=105, value=values).value
        date_number_document(colum)

        return values, document, True
    else:
        values = colum.cell(row=18, column=89).value
        values = colum.cell(row=26, column=105, value=values).value
        machine1 = base["machine2"]
        machine1_driver = machine1["driver"]
        machine1_driver = colum.cell(row=49, column=6, value=machine1_driver).value
        machine1_driver = colum.cell(row=33, column=94, value=machine1_driver).value
        machine1_driver = colum.cell(row=72, column=17, value=machine1_driver).value
        machine1_driver = colum.cell(row=69, column=73, value=machine1_driver).value
        machine1_certificate = machine1['certificate']
        machine1_certificate = colum.cell(row=49, column=51, value=machine1_certificate).value
        machine1_state_signs = machine1['state_signs']
        machine1_state_signs = colum.cell(row=45, column=63, value=machine1_state_signs).value
        machine1_trailer = machine1["trailer"]
        machine1_trailer = colum.cell(row=55, column=83, value=machine1_trailer).value
        document = colum.cell(row=62, column=31, value=document).value
        date_number_document(colum)

    return machine1_driver, machine1_certificate, machine1_state_signs, machine1_trailer, values, document


def count_files_pathlib(folder_path):
    folder = Path(folder_path)

    if not folder.exists():
        return 0

    files = [item.name for item in folder.iterdir() if item.is_file() and item.name != '.DS_Store']
    return len(files)


def distribute_cement(total_weight):
    machine1 = []
    machine2 = []
    remaining_weight = total_weight

    while remaining_weight > 0:
        # Определяем вес для текущего рейса (случайное число от 12.8 до 18.8 или остаток)
        if remaining_weight >= 12.8:
            # Генерируем случайный вес от 12.8 до 18.8, но не больше оставшегося
            load = round(random.uniform(12.8, min(18.8, remaining_weight)), 1)
            print(f"{load}_____________")
        else:
            load = remaining_weight  # последний рейс берет остаток (может быть < 12.8)
            print(load)

        # Распределяем между машинами по очереди
        if len(machine1) <= len(machine2):
            machine1.append(load)
        else:
            machine2.append(load)

        remaining_weight -= load
        remaining_weight = round(remaining_weight, 1)  # округление для избежания ошибок float

    # Проверка общего веса
    sum1 = round(sum(machine1), 1)
    sum2 = round(sum(machine2), 1)
    assert abs((sum1 + sum2) - total_weight) < 0.1, f"Ошибка: общий вес не совпадает! {sum1 + sum2} != {total_weight}"

    return machine1, machine2


number = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
count = int(len(number))
print(len(number), '____')


def number_check(check):
    if check % 2 == 0:
        print(f"число {check} - четное ")
        return 2
    else:
        print(f"число {check} - нечетное ")
        return 1


def test_number_enumerate(list, check):
    number_list = {}

    start = number_check(check)
    for  i, weight in enumerate(list,start=start):

        number_list[check+(i-1)*2] =weight
    return number_list
number = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]

print(f'test-1', test_number_enumerate(number, 1))
number_ = ws.cell(row=4, column=103).value
number = int(number_)

print(number_check(number))

dict_ = {
    1: ['user1', "password", 'login'],
    2: ['user2', "password", 'login'],
    3: ['user3', "password", 'login'],
    4: ['user4', "password", 'login'],
    5: ['user5', "password", 'login']
}
#


file_count = count_files_pathlib(folder_path_machine1)
file_count2 = count_files_pathlib(folder_path_machine2)

print(file_count)
print(file_count2)

print(dict_.get(3))

# print(distribute_cement(319.3))
# print(date_number_document(ws))
