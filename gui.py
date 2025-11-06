import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from pathlib import Path
from functions import distribute_cement, ttn_save_machine1, ttn_save_machine2
from settings import BASE_DIR, base, month
from openpyxl import load_workbook
import calendar


class TTNGeneratorGUI:

    def __init__(self, root):
        self.root = root
        self.root.title("Генератор ТТН накладных")
        self.root.geometry("900x700")

        self.template_path = None
        self.create_widgets()

    def create_widgets(self):
        # Основной фрейм
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Заголовок
        title_label = ttk.Label(main_frame, text="Генератор ТТН накладных",
                                font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=10)

        # Выбор файла ТТН
        ttk.Label(main_frame, text="Файл ТТН шаблона:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.file_entry = ttk.Entry(main_frame, width=60)
        self.file_entry.grid(row=1, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
        ttk.Button(main_frame, text="Обзор", command=self.browse_file).grid(row=1, column=2, padx=5)

        # Общий вес
        ttk.Label(main_frame, text="Общий вес цемента (т):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.weight_var = tk.StringVar()
        weight_entry = ttk.Entry(main_frame, textvariable=self.weight_var, width=20)
        weight_entry.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)

        # Папка вывода
        ttk.Label(main_frame, text="Папка для сохранения:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.output_var = tk.StringVar()
        self.output_var.set(str(BASE_DIR / "export"))
        output_entry = ttk.Entry(main_frame, textvariable=self.output_var, width=60)
        output_entry.grid(row=3, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
        ttk.Button(main_frame, text="Обзор", command=self.browse_output).grid(row=3, column=2, padx=5)

        # Кнопки
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=3, pady=20)

        ttk.Button(button_frame, text="Загрузить файл", command=self.load_file).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Получить вес", command=self.get_weight).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Сгенерировать ТТН", command=self.generate_ttn,
                   style="Accent.TButton").pack(side=tk.LEFT, padx=5)

        # Область вывода
        ttk.Label(main_frame, text="Результаты:").grid(row=5, column=0, sticky=tk.W, pady=5)

        self.output_text = scrolledtext.ScrolledText(main_frame, width=80, height=20,
                                                     font=("Consolas", 10))
        self.output_text.grid(row=6, column=0, columnspan=3, pady=10, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Статус бар
        self.status_var = tk.StringVar()
        self.status_var.set("Готов к работе")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=1, column=0, sticky=(tk.W, tk.E))

        # Настройка расширения
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(6, weight=1)

        # Стиль для акцентной кнопки
        style = ttk.Style()
        style.configure("Accent.TButton", foreground="white", background="#007acc")

    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Выберите файл ТТН",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, filename)

    def browse_output(self):
        directory = filedialog.askdirectory(title="Выберите папку для сохранения")
        if directory:
            self.output_var.set(directory)

    def load_file(self):
        filename = self.file_entry.get()
        if not filename:
            messagebox.showerror("Ошибка", "Выберите файл ТТН")
            return

        try:
            self.template_path = filename
            self.status_var.set(f"Файл загружен: {Path(filename).name}")
            messagebox.showinfo("Успех", "Файл успешно загружен!")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить файл: {str(e)}")

    def get_weight(self):
        if not self.template_path:
            messagebox.showerror("Ошибка", "Сначала загрузите файл ТТН")
            return

        try:
            ttn = load_workbook(self.template_path)
            sheet = ttn.active
            values = sheet.cell(row=16, column=30).value
            self.weight_var.set(str(values) if values else "")
            self.status_var.set(f"Получен вес: {values} т")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось получить вес: {str(e)}")

    def generate_ttn(self):
        if not self.template_path:
            messagebox.showerror("Ошибка", "Сначала загрузите файл ТТН")
            return

        weight_str = self.weight_var.get()
        if not weight_str:
            messagebox.showerror("Ошибка", "Введите общий вес цемента")
            return

        try:
            total_weight = float(weight_str)
            output_dir = Path(self.output_var.get())

            # Создаем папки для сохранения
            (output_dir / "machine1").mkdir(parents=True, exist_ok=True)
            (output_dir / "machine2").mkdir(parents=True, exist_ok=True)

            self.status_var.set("Распределение веса...")
            self.root.update()

            # Очищаем номер документа перед началом
            # global number_document
            # number_document.clear()

            # Распределяем вес
            machine1, machine2 = distribute_cement(total_weight)

            self.status_var.set("Сохранение файлов для machine1...")
            self.root.update()

            # Сохраняем файлы для machine1 с ИСХОДНЫМ шаблоном
            ttn_save_machine1(machine1, self.template_path, output_dir)

            self.status_var.set("Сохранение файлов для machine2...")
            self.root.update()

            # Сохраняем файлы для machine2 с ИСХОДНЫМ шаблоном
            ttn_save_machine2(machine2, self.template_path, output_dir)

            # Выводим результаты
            result = f"""РАСПРЕДЕЛЕНИЕ ЦЕМЕНТА

Общий вес: {total_weight} т

Машина 1 ({base['machine1']['driver']}):
Количество рейсов: {len(machine1)}
Веса по рейсам: {machine1}
Общий вес машины 1: {sum(machine1):.1f} т

Машина 2 ({base['machine2']['driver']}):
Количество рейсов: {len(machine2)}
Веса по рейсам: {machine2}
Общий вес машины 2: {sum(machine2):.1f} т

Проверка: {sum(machine1) + sum(machine2):.1f} т = {total_weight} т

Файлы сохранены в:
{output_dir / 'machine1'}
{output_dir / 'machine2'}
"""
            self.output_text.delete(1.0, tk.END)
            self.output_text.insert(tk.END, result)

            self.status_var.set("Генерация завершена успешно!")
            messagebox.showinfo("Успех", "ТТН накладные успешно сгенерированы!")

        except Exception as e:
            self.status_var.set("Ошибка")
            messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")
            self.output_text.insert(tk.END, f"ОШИБКА: {str(e)}\n")


def main():
    root = tk.Tk()
    app = TTNGeneratorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()


