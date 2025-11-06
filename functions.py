import random


from settings import BASE_DIR, base, month, month__, number_check, examination_in_month_day
from openpyxl import load_workbook
#

def date_number_document(colum):
    # Получаем данные из ячеек
    data_number = colum.cell(row=39, column=8).value
    data_number = int(str(data_number).strip('"'))  # Более безопасное преобразование
    data_month = colum.cell(row=39, column=13).value
    year = colum.cell(row=39, column=28).value
    year = str(year)[0:4]  # Преобразуем в строку и берем первые 4 символа

    # print(f"Начальные данные: число={data_number}, месяц={data_month}, год={year}")

    # count = 3
    current_day = data_number
    current_month = data_month
    current_year = int(year)

    # for i in range(count):
    # Проверяем, не превышает ли текущий день количество дней в месяце
    days_in_month = examination_in_month_day(current_year, current_month)

    if current_day > days_in_month:
        # Переходим к следующему месяцу
        current_day = 1

        # Находим следующий месяц
        for key, value in month.items():
            if value == current_month:
                next_key = int(key) % 12 + 1
                current_month = month[str(next_key)]

                # Если перешли на январь, увеличиваем год
                if next_key == 1:
                    current_year += 1
                break
    else:
        current_day += 1

    # Записываем данные в ячейки
    current_day = colum.cell(row=39, column=8, value=current_day).value
    current_month = colum.cell(row=39, column=13, value=current_month).value
    current_year = colum.cell(row=39, column=28, value=str(current_year)).value
    new_date = []

    for kay, value in month__.items():

        if current_month == value:
            new_date.append(str(current_day))
            new_date.append(kay)
            new_date.append(str(current_year))

    new_date = ".".join(new_date)
    date = colum.cell(row=5, column=103, value=str(new_date)).value
    print(f"день: {current_day},  Месяц: {current_month}, год: {current_year} ----> Дата в ячейке R5cC103 {date}")
    # print(f"Дата добавлена {current_day}:{current_month}:{current_year}")
    # return f"день: {current_day},  Месяц: {current_month}, год: {current_year} ----> Дата в ячейке R5cC103 {date}"


def examination_machine1(colum, row: int, column: int, basa):
    documen_namber = colum.cell(row=4, column=103).value
    documen_date = colum.cell(row=5, column=103).value
    document = f"ТТН {documen_namber} от {documen_date}"
    values = colum.cell(row=row, column=column).value
    if values == "Джанбулатов Руслан Джанболатович":
        ducument = colum.cell(row=62, column=31, value=document).value
        values = colum.cell(row=18, column=89).value
        values = colum.cell(row=26, column=105, value=values).value
        date_number_document(colum)

        return values, ducument, True
    else:
        values = colum.cell(row=18, column=89).value
        values = colum.cell(row=26, column=105, value=values).value
        machine1 = basa["machine1"]
        machine1_driver = machine1["driver"]
        machine1_driver = colum.cell(row=49, column=6, value=machine1_driver).value
        machine1_driver = colum.cell(row=33, column=94, value=machine1_driver).value
        machine1_driver = colum.cell(row=72, column=17, value=machine1_driver).value
        machine1_driver = colum.cell(row=69, column=73, value=machine1_driver).value
        machine1_certificate = machine1['certificate']
        machine1_certificate = colum.cell(row=49, column=51, value=machine1_certificate).value
        machine1_state_signs = machine1['state_signs']
        machine1_state_signs = colum.cell(row=45, column=63, value=machine1_state_signs).value
        machine1_trailer = machine1["trailer"]
        machine1_trailer = colum.cell(row=55, column=83, value=machine1_trailer).value
        ducument = colum.cell(row=62, column=31, value=document).value

    return machine1_driver, machine1_certificate, machine1_state_signs, machine1_trailer, values, ducument


def examination_machine2(colum, row: int, column: int, basa):
    document_number = colum.cell(row=4, column=103).value
    document_date = colum.cell(row=5, column=103).value
    document = f"ТТН {document_number} от {document_date}"
    # print(document)
    values = colum.cell(row=row, column=column).value
    if values == "Гайирбеков Ибрагим Расулович":
        document = colum.cell(row=62, column=31, value=document).value
        values = colum.cell(row=18, column=89).value
        values = colum.cell(row=26, column=105, value=values).value
        date_number_document(colum)

        return values, document, True
    else:
        values = colum.cell(row=18, column=89).value
        values = colum.cell(row=26, column=105, value=values).value
        machine1 = base["machine2"]
        machine1_driver = machine1["driver"]
        machine1_driver = colum.cell(row=49, column=6, value=machine1_driver).value
        machine1_driver = colum.cell(row=33, column=94, value=machine1_driver).value
        machine1_driver = colum.cell(row=72, column=17, value=machine1_driver).value
        machine1_driver = colum.cell(row=69, column=73, value=machine1_driver).value
        machine1_certificate = machine1['certificate']
        machine1_certificate = colum.cell(row=49, column=51, value=machine1_certificate).value
        machine1_state_signs = machine1['state_signs']
        machine1_state_signs = colum.cell(row=45, column=63, value=machine1_state_signs).value
        machine1_trailer = machine1["trailer"]
        machine1_trailer = colum.cell(row=55, column=83, value=machine1_trailer).value
        document = colum.cell(row=62, column=31, value=document).value
        # date_number_document(colum)

    return machine1_driver, machine1_certificate, machine1_state_signs, machine1_trailer, values, document


def distribute_cement(total_weight):
    machine1 = []
    machine2 = []
    remaining_weight = total_weight

    while remaining_weight > 0:
        # Определяем вес для текущего рейса (случайное число от 12.8 до 18.8 или остаток)
        if remaining_weight >= 12.8:
            # Генерируем случайный вес от 12.8 до 18.8, но не больше оставшегося
            load = round(random.uniform(12.8, min(18.8, remaining_weight)), 1)
        else:
            load = remaining_weight  # последний рейс берет остаток (может быть < 12.8)

        # Распределяем между машинами по очереди
        if len(machine1) <= len(machine2):
            machine1.append(load)
        else:
            machine2.append(load)

        remaining_weight -= load
        remaining_weight = round(remaining_weight, 1)  # округление для избежания ошибок float

    # Проверка общего веса
    sum1 = round(sum(machine1), 1)
    sum2 = round(sum(machine2), 1)
    assert abs((sum1 + sum2) - total_weight) < 0.1, f"Ошибка: общий вес не совпадает! {sum1 + sum2} != {total_weight}"
    print(f"Количество рейсов первой машины:{len(machine1)} количество рейсов второй машины: {len(machine2)}")
    return machine1, machine2



def ttn_save_machine1(machine1, template_path, output_dir ):
    try:
        # Загружаем ЧИСТЫЙ шаблон для machine1
        workbook = load_workbook(template_path)
        column = workbook.active

        number_ = int(column.cell(row=4, column=103).value)
        start = number_check(number_)  # проверка

        for i, weight in enumerate(machine1, start=start):
            number = number_ + (i - start) * 2  # решение
            column.cell(row=4, column=103, value=number)
            column.cell(row=16, column=30, value=weight)
            column.cell(row=17, column=30, value=weight)
            column.cell(row=18, column=30, value=weight)
            summ = column.cell(row=16, column=30).value * column.cell(row=16, column=37).value
            column.cell(row=17, column=89, value=summ)
            column.cell(row=18, column=89, value=summ)
            column.cell(row=16, column=89, value=summ)
            date_document = column.cell(row=5, column=103).value
            examination_machine1(column, 49, 6, base)

            workbook.save(
                output_dir / "machine1" / f"Tovarno-transportnaya nakladnaya № {number} ot {date_document}.xlsx")

    except Exception as e:
        return f'Ошибка:{e}'


def ttn_save_machine2(machine2, template_path, output_dir):
    try:
        # Загружаем ЧИСТЫЙ шаблон для machine2 (исходный файл)
        workbook = load_workbook(template_path)
        column = workbook.active

        number_ = int(column.cell(row=4, column=103).value)+ 1  # взял число из списка
        start = number_check(number_)  # проверка
        print(number_)

        for i, weight in enumerate(machine2, start=start):
            number = number_ + (i - start) * 2  # решение
            column.cell(row=4, column=103, value=number)
            column.cell(row=16, column=30, value=weight)
            column.cell(row=17, column=30, value=weight)
            column.cell(row=18, column=30, value=weight)
            summ = column.cell(row=16, column=30).value * column.cell(row=16, column=37).value
            column.cell(row=16, column=89, value=summ)
            column.cell(row=17, column=89, value=summ)
            column.cell(row=18, column=89, value=summ)
            date_document = column.cell(row=5, column=103).value
            examination_machine2(column, 49, 6, base)
            workbook.save(
                output_dir / "machine2" / f"Tovarno-transportnaya nakladnaya № {number} ot {date_document}.xlsx")

    except Exception as e:
        return f'Ошибка {e}'

def get_weight_from_file(filepath):
    try:
        ttn = load_workbook(filepath)
        sheet = ttn.active
        values = sheet.cell(row=16, column=30).value

        return values
    except Exception as e:
        raise Exception(f"Не удалось получить вес из файла: {str(e)}")
